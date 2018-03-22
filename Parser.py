
##################### JUNTAS #################################

import csv

def agrega_votos_recibidos():
    lista_votos =[]
    leer_votos_x_junta(lista_votos)
    with open('./Documentos/JuntasSinVotos.csv','r') as csvinput, open('./Documentos/Juntas.csv', 'w') as csvoutput:
            writer = csv.writer(csvoutput, lineterminator='\n')
            reader = csv.reader(csvinput)

            all = []

            for row in reader:
                row.append(buscar_votos_x_junta(lista_votos, row[4]))
                all.append(row)

            writer.writerows(all)

def leer_votos_x_junta(lista_resultado):
    with open('./Documentos/JuntaxVRecibidos.csv','r') as csvinput:
        reader = csv.reader(csvinput)
        for row in reader:
            lista_resultado.append(row)

def buscar_votos_x_junta(lista_busqueda, key):
    for i in lista_busqueda:
        if(i[0] == key):
            return i[1]

def buscar_faltantes():
    cont = 1
    with open('./Documentos/JuntasSinVotos.csv','r') as csvinput:
        reader = csv.reader(csvinput)

        for row in reader:
            if (int(row[4]) != cont):
                print(cont, row[4])
            cont = int(row[4])+1

def buscar_errores():
    with open('./Documentos/JuntasSinVotos.csv','r') as csvinput:
        reader = csv.reader(csvinput)

        for row in reader:
            if (row[0].isdigit() or row[1].isdigit() or row[2].isdigit()):
                print(row[4])

################## INDICADORES ##########################
import os
from openpyxl import load_workbook
import unicodedata

def convertir_archivo_indicadores():
    os.chdir("C:/Users/Mauro/Desktop/IA/PC1")
    book = load_workbook('./Documentos/indicadores.xlsx')
    sheets = book.sheetnames
##    with open('./Documentos/Indicadores_x_Canton.csv', 'w') as csvoutput:
##        writer = csv.writer(csvoutput, lineterminator='\n')
##        for i in sheets:
##            sheet = book[i]
##            cells = sheet['H2:BO2'][0]
##            for c in cells:
##                if(c.value!=None):
##                    canton = c.value.upper()
##                    canton = remove_accents(canton)
##                    row = [sheet.title, canton]
##                    writer.writerow(row)
    
    cantones = []
    rows = [5,6,7,8,0,0,9,0,0,10,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,
            0,0,0,0,0,0,0,0,14,15,16,0,0,17,0,0,22,24,25,26,28,29,30,31,32,33,34,0,0,0,0,0,0,
            38,0,0,0,0,0,39,41,42,0,0,0,43,47,0,0,48,0,0,49,0,0,0,0,0,50,51,0,0,0,0,0,0,0]
    for i in range(7):
        sheet = book[sheets[i]]
        c = 9
        while(True):
            if(sheet.cell(row = 5, column = c).value == None):
                break
            canton = []
            for r in rows:        
                if r == 0:
                    canton.append("")
                else:
                    canton.append(sheet.cell(row = r, column = c).value)
            
            canton[4], canton[5] = proporciones(canton[0], canton[3])
            #hombres - mujeres
            canton[7], canton[8] = (100*canton[0])/(canton[6]+100), canton[0] - ((100*canton[0])/(canton[6]+100))
            canton[43], canton[44] = proporciones(canton[40], canton[42])
            canton[46], canton[47] = proporciones(canton[40], canton[45])
            canton[79], canton[80] = proporciones(canton[0], canton[78])
            canton[82], canton[83] = proporciones(canton[0], canton[81])
            canton[85], canton[86] = proporciones(canton[0], canton[84])
            cont = sheet.cell(row = 5, column = c).value
            cantones.append(canton)
            c+=3
    rellenar_edad(cantones, book[sheets[7]])
    rellenar_escolaridad(cantones, book[sheets[8]])
    rellenar_fuera_fuerza_laboral(cantones, book[sheets[9]])
    rellenar_sector(cantones, book[sheets[10]])
    rellenar_seguro(cantones, book[sheets[11]])
    rellenar_tics(cantones, book[sheets[12]])

    with open('./Documentos/Indicadores_x_Canton.csv','r') as csvinput, open('./Documentos/Indicadores_x_Canton_Completo.csv', 'w') as csvoutput:
            writer = csv.writer(csvoutput, lineterminator='\n')
            reader = csv.reader(csvinput)

            all = []
            index = 0
            for row in reader:
                all.append(row+cantones[index])
                index+=1
            
            writer.writerows(all)


def remove_accents(input_str):
    nfkd_form = unicodedata.normalize('NFKD', input_str)
    return u"".join([c for c in nfkd_form if not unicodedata.combining(c)])

def proporciones(total, porcentaje):
    calculado = porcentaje/100 *total
    return calculado, total-calculado

def rellenar_edad(cantones,sheet):
    c = 6
    for canton in cantones:
        index = 10
        for edad in range(3, 18):
            canton[index] = sheet.cell(row = edad, column = c).value
            index+=1
        index = 25
        for edad in range(3, 18):
            canton[index] = abs(sheet.cell(row = edad, column = c+1).value)
            index+=1
        c+=2

def rellenar_escolaridad(cantones, sheet):
    c = 2
    for canton in cantones:
        index = 59
        for escolaridad in range(2, 8):
            canton[index] = sheet.cell(row = escolaridad, column = c).value
            index+=1
        c+=1

def rellenar_fuera_fuerza_laboral(cantones, sheet):
    c = 2
    for canton in cantones:
        index = 66
        for fuerza in range(2, 7):
            canton[index] = sheet.cell(row = fuerza, column = c).value
            index+=1
        c+=1

def rellenar_sector(cantones, sheet):
    c = 2
    for canton in cantones:
        index = 74
        for sector in range(2, 5):
            canton[index] = sheet.cell(row = sector, column = c).value
            index+=1
        c+=1

def rellenar_seguro(cantones, sheet):
    c = 3
    sector = 1
    for canton in cantones:
        for i in range(4):
            if sheet.cell(row = sector+i, column = 2).value == "Otras formas":
                canton[89] = sheet.cell(row = sector+i, column = c).value
            elif sheet.cell(row = sector+i, column = 2).value == "Indirecto":
                canton[88] = sheet.cell(row = sector+i, column = c).value
            elif sheet.cell(row = sector+i, column = 2).value == "Directo":
                canton[87] = sheet.cell(row = sector+i, column = c).value                         
        sector+=8

def rellenar_tics(cantones, sheet):
    c = 3
    tics = 1
    for canton in cantones:
        for i in range(8):
            if sheet.cell(row = tics+i, column = 2).value == "Agua":
                canton[98] = sheet.cell(row = tics+i, column = c).value
            elif sheet.cell(row = tics+i, column = 2).value == "Servicio sanitario":
                canton[97] = sheet.cell(row = tics+i, column = c).value
            elif sheet.cell(row = tics+i, column = 2).value == "Electricidad":
                canton[96] = sheet.cell(row = tics+i, column = c).value
            elif sheet.cell(row = tics+i, column = 2).value == "Internet":
                canton[95] = sheet.cell(row = tics+i, column = c).value
            elif sheet.cell(row = tics+i, column = 2).value == "Computadora":
                canton[94] = sheet.cell(row = tics+i, column = c).value
            elif sheet.cell(row = tics+i, column = 2).value == "Teléfono residencial":
                canton[93] = sheet.cell(row = tics+i, column = c).value
            elif sheet.cell(row = tics+i, column = 2).value == "Teléfono celular":
                canton[92] = sheet.cell(row = tics+i, column = c).value
        tics+=11


#convertir_archivo_indicadores()
    

 
        


    
